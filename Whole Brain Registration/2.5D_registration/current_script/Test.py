import argparse
import os
import sys
import copy
sys.path.append("")
from current_script.script_util import (
    add_dict_to_argparser,
    parse_resume_step_from_filename, set_random_seed, args_to_dict,
)
from util import logger
from current_script.model_util import voxelmorph_defaults
import torch
import json
import time
from util.data_util.dataset import TrainPairDataset, ValPairDataset
from torch.utils.data import DataLoader
from util.loss.NCCLoss import NCCLoss
from util.loss.MSELoss import MSELoss
from util.loss.DiceLoss import DiceLoss
from util.loss.GradientLoss2D import GradientLoss2D
from model.registration.voxelmorph.layers import SpatialTransformer
from util.metric.registration_metric import folds_percent_metric, ssim_metric, jacobian_determinant
from util.metric.segmentation_metric import dice_metric3
from util.ModelSaver import ModelSaver
from shutil import copyfile
from tqdm import tqdm
from torch.autograd import Variable
# from util.visual.visual1.visual_image import cat_imgs, tensor2img, save_img
from util.visual.image_util import save_deformation_figure, save_det_figure, save_dvf_figure
from model.registration.voxelmorph.voxelmorph import VxmDense
import monai
from torch.utils.data import Dataset
import numpy as np
from util.data_util.io_util import read_3d_data
import SimpleITK as sitk
import toml

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

region_list = ["BS", "CB", "CP", "CTX", "HPF", "LGvc", "lgd", "RN", "MH_LH", "fr", "aco_act"]
region_dict={
    'BS': 1,
    'CB': 2,
    'CP': 3,
    'CTX': 4,
    'HPF':5, 
    
    "LGvc": 6,
    "lgd": 7,
    "RN": 8,
    "MH_LH": 9,
    "fr": 10,
    "aco_act": 11,
}
def save_as_xlsx_multiple_rows(res, file_path, name, metrics, number=0):
    """
    将形如:
        res = {
            'region1': [val1, val2, val3, ...],
            'region2': [val1, val2, val3, ...]
            ...
        }
    的字典写成 XLSX：
    
    行 1：各 region 分段合并单元格，显示 region 名称
    行 2：各 region 的列，分别显示传入的指标名称
    行 3：与行 2 对应的数值
    支持动态指标数量。
    
    参数:
        res: 字典，包含 region 和对应的指标值
        file_path: 保存的文件路径
        name: 当前数据的名称
        metrics: 指标名称列表，例如 ["Dice", "Affine intensity", "SyN intensity"]
        number: 起始行偏移量
    """
    if not os.path.exists(file_path) or number == 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # 依次写各 region，每个 region 占 len(metrics) 列
        current_index = 0
        ws.cell(row=3, column=1, value=name)
        for region, values in res.items():
            start_col = 2 + current_index * len(metrics)
            end_col = start_col + len(metrics) - 1
            
            # (行 1) 合并单元格，并写 region 名称
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )
            ws.cell(row=1, column=start_col, value=region)
            
            # (行 2) 写指标名称
            for i, metric in enumerate(metrics):
                ws.cell(row=2, column=start_col + i, value=metric)
            
            # (行 3) 写具体数值
            for i, value in enumerate(values):
                ws.cell(row=3, column=start_col + i, value=value)
            
            current_index += 1
    
    else:
        wb = load_workbook(file_path)
        ws = wb.active
        
        st = number + 3
        
        current_index = 0

        ws.cell(row=st, column=1, value=name)
        for region, values in res.items():
            start_col = 2 + current_index * len(metrics)
            
            # (行 3) 写具体数值
            for i, value in enumerate(values):
                ws.cell(row=st, column=start_col + i, value=value)
            
            current_index += 1
        
    # 最后保存 XLSX 到指定目录
    wb.save(file_path)


def create_model_var1(
        spatial_dims,
        int_steps,
        use_probs,
        src_feats,
        trg_feats
):
    enc_nf = [16, 32, 32, 32]
    dec_nf = [32, 32, 32, 32, 32, 16, 16]
    return VxmDense(
        spatial_dims=spatial_dims,
        nb_unet_features=[enc_nf, dec_nf],
        int_steps=int_steps,
        use_probs=use_probs,
        src_feats=src_feats,
        trg_feats=trg_feats,
    )


def voxelmorph_defaults_var1():
    config = dict(
        spatial_dims=2,
        int_steps=0,
        use_probs=False,
        src_feats=2,
        trg_feats=2,
    )
    return config

def voxelmorph_defaults_var2():
    config = dict(
        spatial_dims=2,
        int_steps=0,
        use_probs=False,
        src_feats=1,
        trg_feats=1,
    )
    return config


def main():
    set_random_seed(seed=0)
    args = create_argparser().parse_args()
    if len(args.resolution) > 0:
        args.resolution = tuple([int(i) for i in args.resolution.split(',')])
    os.environ['LOGDIR'] = args.logdir
    os.environ['CUDA_VISIBLE_DEVICES'] = args.gpu
    os.environ['OPENAI_LOG_FORMAT'] = 'stdout,log,csv,tensorboard'
    logger.configure()
    logger.log(list(vars(args).items()))
    logger.log(list(voxelmorph_defaults_var1().items()))
    logger.log(f"start train time: {time.asctime()}")
    train_path = os.path.abspath(__file__)
    logger.log(f"train file path: {train_path}")
    copyfile(train_path, os.path.join(logger.get_dir(), os.path.basename(train_path)))

    if args.is_sup:
        model_1 = create_model_var1(**voxelmorph_defaults_var1())
        model_2 = create_model_var1(**voxelmorph_defaults_var1())
    else:
        model_1 = create_model_var1(**voxelmorph_defaults_var2())
        model_2 = create_model_var1(**voxelmorph_defaults_var2())

    total_params = sum([param.nelement() for param in model_1.parameters()])
    logger.log("Number of parameter: %.2fM" % (total_params / 1e6))

    optimizer_1 = torch.optim.Adam(model_1.parameters(), lr=args.lr)
    optimizer_2 = torch.optim.Adam(model_2.parameters(), lr=args.lr)
    device = torch.device("cuda:0" if len(args.gpu) > 0 else "cpu")
    resume_epoch = 0
    if len(args.resume_checkpoint) > 0:
        logger.log(f"loading model from {args.resume_checkpoint}")
        resume_epoch = parse_resume_step_from_filename(args.resume_checkpoint)
        state_dict = torch.load(args.resume_checkpoint, map_location=device)
        model_1.load_state_dict(state_dict['model_1'])
        optimizer_1.load_state_dict(state_dict['optimizer_1'])

        # for state in optimizer_1.state.values():
        #     for k, v in state.items():
        #         if torch.is_tensor(v):
        #             state[k] = v.cuda()

        # model_2.load_state_dict(state_dict['model_2'])
        # optimizer_2.load_state_dict(state_dict['optimizer_2'])

        # for state in optimizer_2.state.values():
        #     for k, v in state.items():
        #         if torch.is_tensor(v):
        #             state[k] = v.cuda()

    model_1.to(device)
    model_2.to(device)
    gpu_num = len(args.gpu.split(","))

    if gpu_num > 1:
        model_1 = torch.nn.DataParallel(model_1, device_ids=[i for i in range(gpu_num)])
        model_2 = torch.nn.DataParallel(model_2, device_ids=[i for i in range(gpu_num)])
    model_without_ddp_1 = model_1.module if gpu_num > 1 else model_1
    model_without_ddp_2 = model_2.module if gpu_num > 1 else model_2

    
    if args.data_dir.endswith('toml'):
        with open(args.data_dir, 'r') as f:
            dataset_config = toml.load(f)
    elif args.data_dir.endswith('json'):
        with open(args.data_dir, 'r') as f:
            dataset_config = json.load(f)
    else:
        raise Exception("dataset config file format error")

    num_classes = args.num_classes

    transform = None
    if len(args.resolution) > 0:
        transform = monai.transforms.ResizeD(spatial_size=args.resolution,
                                             keys=['volume1', 'volume2', 'label1', 'label2'],
                                             mode=['trilinear', 'trilinear', 'nearest', 'nearest'])
    # train_dataset = TrainPairDataset(dataset_config, dataset_type='train',
    #                                  registration_type=args.registration_type,
    #                                  transform=transform)

    val_dataset = TrainPairDataset(dataset_config, dataset_type='test',
                                 registration_type=args.registration_type,
                                 transform=transform)
    # logger.log(f'train dataset size: {len(train_dataset)}')
    logger.log(f'val dataset size: {len(val_dataset)}')
    # train_dataloader = DataLoader(train_dataset, batch_size=1)
    val_dataloader = DataLoader(val_dataset, batch_size=1, shuffle=False)
    # modelsaver = ModelSaver(args.max_save_num)
    run_test(args, model_1, model_2, val_dataloader, device, num_classes)
    logger.log(f"end train time: {time.asctime()}")


def extract_slices(img, i, batch_size, dim):
    if dim == 2:
        return img[:, :, i:i + batch_size, :, :]
    elif dim == 3:
        return img[:, :, :, i:i + batch_size, :]
    elif dim == 4:
        return img[:, :, :, :, i:i + batch_size]
    else:
        raise Exception("dim error")


def extract_reg_slices(volume1, volume2, label1, label2, i, batch_size, dim):
    img1 = extract_slices(volume1, i, batch_size, dim)
    img2 = extract_slices(volume2, i, batch_size, dim)
    mask1 = extract_slices(label1, i, batch_size, dim) if label1 != [] else []
    mask2 = extract_slices(label2, i, batch_size, dim) if label2 != [] else []
    return img1, img2, mask1, mask2


def permute_img(img, B, C, H, W, D, dim):
    if dim == 2:
        return img.permute((0, 2, 1, 3, 4)).reshape((B * H, C, W, D)).contiguous()
    elif dim == 3:
        return img.permute((0, 3, 1, 2, 4)).reshape((B * W, C, H, D)).contiguous()
    elif dim == 4:
        return img.permute((0, 4, 1, 2, 3)).reshape((B * D, C, H, W)).contiguous()
    else:
        raise Exception("dim error")


def permute_reg_imgs(img1, img2, mask1, mask2, B, C, H, W, D, dim):
    img1 = permute_img(img1, B, C, H, W, D, dim)
    img2 = permute_img(img2, B, C, H, W, D, dim)
    mask1 = permute_img(mask1, B, C, H, W, D, dim) if mask1 != [] else []
    mask2 = permute_img(mask2, B, C, H, W, D, dim) if mask2 != [] else []
    return img1, img2, mask1, mask2


def unpermute_img(img, B, C, H, W, D, dim):
    if img == []:
        return img

    if dim == 2:
        return img.reshape((B, H, C, W, D)).permute((0, 2, 1, 3, 4)).contiguous()
    elif dim == 3:
        return img.reshape((B, W, C, H, D)).permute((0, 2, 3, 1, 4)).contiguous()
    elif dim == 4:
        return img.reshape((B, D, C, H, W)).permute((0, 2, 3, 4, 1)).contiguous()
    else:
        raise Exception("dim error")


def concat_dvf_3d(dvf, dim):
    remaining_channel = torch.zeros((dvf.shape[0], 1) + tuple(dvf.shape[2:]), device=dvf.device).float()
    dvf = torch.cat((dvf, remaining_channel), dim=1)
    if dim == 2:
        dvf = dvf[:, [2, 0, 1], ...]
    elif dim == 3:
        dvf = dvf[:, [0, 2, 1], ...]
    return dvf


def run_test(args, model_1, model_2, val_dataloader, device, num_classes):
    model_1.eval()
    model_2.eval()
    batch_size = args.batch_size
    view_1_dim = args.view_1_dim
    view_2_dim = args.view_2_dim
    # TEST
    if not args.use_view_1:
        view_1_dim, view_2_dim = view_2_dim, view_1_dim
        model_1, model_2 = model_2, model_1
    
    with torch.no_grad():
        from collections import defaultdict
        res = {key: [0] for key in region_dict.keys()}
        count = 0
        for img_dict in tqdm(val_dataloader):
            volume1, volume2 = img_dict['volume1'], img_dict['volume2']
            label1, label2 = img_dict['label1'], img_dict['label2']
            id1, id2 = img_dict['id1'][0], img_dict['id2'][0]
            # TEST
            volume1 = volume1.to(device)
            label1 = label1.to(device) if label1 != [] else label1

            volume2 = volume2.to(device)
            label2 = label2.to(device) if label2 != [] else label2
            
            label_test = img_dict["label_test"]
            label_test = label_test.to(device) if label_test != [] else label_test

            view_1_size = volume1.shape[view_1_dim]
            view_1_warp_volume1, view_1_warp_label1, view_1_dvf = [], [], []
            view_1_warp_label1_nearest = []
            for i in range(0, view_1_size, batch_size):
                img1, img2, mask1, mask2 = extract_reg_slices(volume1, volume2, label1, label2,
                                                              i, batch_size, view_1_dim)
                B, C, H, W, D = img1.shape
                img1, img2, mask1, mask2 = permute_reg_imgs(img1, img2, mask1, mask2, B, C, H, W, D, view_1_dim)
                
                mask_test = extract_slices(label_test, i, batch_size, view_1_dim)
                mask_test = permute_img(mask_test, B, C, H, W, D, view_1_dim)
                

                # TEST 只输入主脑区
                mask1_with_main_label = mask1.clone()
                # mask1_with_main_label[mask1_with_main_label > 5] = 1
                mask2_with_main_label = mask2.clone()
                # mask2_with_main_label[mask2_with_main_label > 5] = 1
                
                warp_img1, warp_mask1, dvf_i = val_one_view(args, model_1, img1, img2, mask1_with_main_label, mask2_with_main_label, "val_view_1_")

                if mask1 != []:
                    warp_mask1_nearest = SpatialTransformer(mode='nearest')(mask_test.float(), dvf_i)
                    warp_mask1_nearest = unpermute_img(warp_mask1_nearest, B, C, H, W, D, view_1_dim)
                    view_1_warp_label1_nearest.append(warp_mask1_nearest)

                warp_img1 = unpermute_img(warp_img1, B, C, H, W, D, view_1_dim)
                warp_mask1 = unpermute_img(warp_mask1, B, C, H, W, D, view_1_dim)
                dvf_i = unpermute_img(dvf_i, B, dvf_i.shape[1], H, W, D, view_1_dim)

                view_1_warp_volume1.append(warp_img1)
                if warp_mask1 != []:
                    view_1_warp_label1.append(warp_mask1)
                view_1_dvf.append(dvf_i)

            view_1_warp_volume1 = torch.cat(view_1_warp_volume1, dim=view_1_dim)
            if view_1_warp_label1 != []:
                view_1_warp_label1 = torch.cat(view_1_warp_label1, dim=view_1_dim)
            view_1_warp_label1_nearest = torch.cat(view_1_warp_label1_nearest, dim=view_1_dim)

            view_1_dvf = torch.cat(view_1_dvf, dim=view_1_dim)
            view_1_dvf = concat_dvf_3d(view_1_dvf, dim=view_1_dim)

            view_2_size = volume1.shape[view_2_dim]
            view_2_warp_volume1, view_2_warp_label1, view_2_dvf = [], [], []
            view_2_warp_label1_nearest = []
            for i in range(0, view_2_size, batch_size):
                img1, img2, mask1, mask2 = extract_reg_slices(volume1, volume2, label1, label2,
                                                              i, batch_size, view_2_dim)
                B, C, H, W, D = img1.shape
                img1, img2, mask1, mask2 = permute_reg_imgs(img1, img2, mask1, mask2, B, C, H, W, D, view_2_dim)

                # TEST 只输入主脑区
                mask1_with_main_label = mask1.clone()
                # mask1_with_main_label[mask1_with_main_label > 5] = 1
                mask2_with_main_label = mask2.clone()
                # mask2_with_main_label[mask2_with_main_label > 5] = 1
                
                warp_img1, warp_mask1, dvf_i = val_one_view(args, model_2, img1, img2, mask1_with_main_label, mask2_with_main_label, "val_view_2_")

                if mask1 != []:
                    warp_mask1_nearest = SpatialTransformer(mode='nearest')(mask1.float(), dvf_i)
                    warp_mask1_nearest = unpermute_img(warp_mask1_nearest, B, C, H, W, D, view_2_dim)
                    view_2_warp_label1_nearest.append(warp_mask1_nearest)

                warp_img1 = unpermute_img(warp_img1, B, C, H, W, D, view_2_dim)
                warp_mask1 = unpermute_img(warp_mask1, B, C, H, W, D, view_2_dim)
                dvf_i = unpermute_img(dvf_i, B, dvf_i.shape[1], H, W, D, view_2_dim)

                view_2_warp_volume1.append(warp_img1)
                if warp_mask1 != []:
                    view_2_warp_label1.append(warp_mask1)
                view_2_dvf.append(dvf_i)

            view_2_warp_volume1 = torch.cat(view_2_warp_volume1, dim=view_2_dim)
            if view_2_warp_label1 != []:
                view_2_warp_label1 = torch.cat(view_2_warp_label1, dim=view_2_dim)
            view_2_warp_label1_nearest = torch.cat(view_2_warp_label1_nearest, dim=view_2_dim)

            view_2_dvf = torch.cat(view_2_dvf, dim=view_2_dim)
            view_2_dvf = concat_dvf_3d(view_2_dvf, dim=view_2_dim)

            # loss_dict = compute_consistency_loss(args, view_1_warp_volume1, view_2_warp_volume1,
            #                                      view_1_warp_label1, view_2_warp_label1)
            # for key, value in loss_dict.items():
            #     logger.logkv_mean("val_" + key, value.item())

            view_1_metric_dict = compute_metric(view_1_dvf, view_1_warp_volume1, view_1_warp_label1_nearest,
                                                volume2, label2, num_classes)
            view_2_metric_dict = compute_metric(view_2_dvf, view_2_warp_volume1, view_2_warp_label1_nearest,
                                                volume2, label2, num_classes)
            
            for i in range(num_classes):
                res[region_list[i]][0] = view_1_metric_dict[f'dice_class_{i}']
            save_as_xlsx_multiple_rows(res, os.path.join(args.logdir, f"res_dice.xlsx"), id1, ["Dice"], count)
            count += 1
            
            save_dir = os.path.join(args.logdir, "warped_data", id1)
            os.makedirs(save_dir, exist_ok=True)
            sitk.WriteImage(sitk.GetImageFromArray(view_1_warp_volume1.cpu().numpy()), os.path.join(save_dir, "warped_volume.nii.gz"))
            sitk.WriteImage(sitk.GetImageFromArray(volume2.cpu().numpy()), os.path.join(save_dir, "fixed.nii.gz"))
            sitk.WriteImage(sitk.GetImageFromArray(view_1_warp_label1_nearest.cpu().numpy()), os.path.join(save_dir, "warped_label.nii.gz"))
            sitk.WriteImage(sitk.GetImageFromArray(label2.cpu().numpy()), os.path.join(save_dir, "fixed_label.nii.gz"))
                
        # import pandas as pd
        # df = pd.DataFrame.from_dict(res, orient='index')
        # # 将 DataFrame 写入 Excel 文件
        # df.to_excel(f"{args.logdir}/res_dice.xlsx", sheet_name='Sheet1', index=True, header=True)
        # logger.dumpkvs()
            


def save_dvf(dvf, output_dir, view_tag):
    save_deformation_figure(output_dir, view_tag + '_deformation', dvf[0].detach().cpu(),
                            grid_spacing=dvf[0].shape[-1] // 50, linewidth=1, color='darkblue')
    save_dvf_figure(output_dir, view_tag + '_dvf', dvf[0].detach().cpu())
    det = jacobian_determinant(dvf[0].detach().cpu())
    save_det_figure(output_dir, view_tag + '_jacobian_det', det, normalize_by='slice', threshold=0,
                    cmap='Blues')
    # save_RGB_dvf_figure(output_dir, view_tag + '_rgb_dvf', dvf[0].detach().cpu())


def save_img_list(img_list1, img_list2, num_classes, output_path, view_dim):
    interval = img_list1[0].shape[view_dim] // 15
    img_list1 = [img.cpu() for img in img_list1]
    img_list2 = [img.cpu() / num_classes for img in img_list2]
    img_list = img_list1 + img_list2
    img = cat_imgs(img_list, interval=interval, view_dim=view_dim)
    img = tensor2img(img, nrow=1, min_max=(0, 1))

    save_img(img, output_path)


def val_one_view(args, model, img1, img2, mask1, mask2, identifier):
    if args.is_sup:
        _img1 = torch.cat([img1, mask1.float()], dim=1)
        _img2 = torch.cat([img2, mask2.float()], dim=1)
    else:
        _img1 = img1
        _img2 = img2

    dvf_i = model(_img1, _img2)

    loss_dict = compute_loss(args, dvf_i, img1, img2, mask1, mask2)
    for key, value in loss_dict.items():
        logger.logkv_mean(identifier + key, value.item())
    warp_img1 = SpatialTransformer(mode='bilinear')(img1, dvf_i)
    warp_mask1 = []
    if mask1 != []:
        warp_mask1 = SpatialTransformer(mode='bilinear')(mask1.float(), dvf_i)
    return warp_img1, warp_mask1, dvf_i


def compute_metric(dvf, warp_volume1, warp_label1, volume2, label2, num_classes):
    metric_dict = dict()
    # metric_dict['dice'] = dice_metric3(warp_label1, label2, num_classes, reduction="mean").item()
    dice_list = dice_metric3(warp_label1, label2, num_classes, reduction=None)
    for i, value in enumerate(dice_list):
        metric_dict[f'dice_class_{i}'] = value.item()
    # metric_dict['ASD'] = ASD_metric2(warp_label1, label2, num_classes).item()
    # metric_dict['HD'] = HD_metric2(warp_label1, label2, num_classes).item()
    metric_dict['SSIM'] = ssim_metric(warp_volume1, volume2).item()
    metric_dict['folds_percent'] = folds_percent_metric(dvf).item()
    # metric_dict['folds_count'] = folds_count_metric(dvf).item()
    # metric_dict['mse'] = mse_metric(warp_volume1, volume2).item()
    return metric_dict


def compute_consistency_loss(args, view_1_warp_volume1, view_2_warp_volume1,
                             view_1_warp_label1, view_2_warp_label1):
    # similarity_loss = NCCLoss(spatial_dims=3, kernel_size=9)(view_1_warp_volume1, view_2_warp_volume1)
    similarity_loss = MSELoss()(view_1_warp_volume1, view_2_warp_volume1)
    seg_loss = torch.tensor(0., device=view_1_warp_volume1.device)
    if view_1_warp_label1 != [] and view_2_warp_label1 != []:
        seg_loss = DiceLoss()(view_1_warp_label1, view_2_warp_label1)
    # similarity_loss = Variable(similarity_loss, requires_grad=True)
    # seg_loss = Variable(seg_loss, requires_grad=True)
    total_loss = similarity_loss + args.seg_loss_weight * seg_loss
    total_loss = Variable(total_loss, requires_grad=True)
    total_loss = args.consistency_loss_weight * total_loss
    loss_dict = {"consistency_total_loss": total_loss,
                 "consistency_similarity_loss": similarity_loss,
                 "consistency_seg_loss": seg_loss}
    return loss_dict


def compute_loss(args, dvf, img1, img2, mask1, mask2):
    warp_img1 = SpatialTransformer(mode='bilinear')(img1, dvf)
    similarity_loss = NCCLoss(spatial_dims=2, kernel_size=9)(warp_img1, img2)
    seg_loss = torch.tensor(0., device=dvf.device)
    if mask1 != [] and mask2 != []:
        num_class = torch.max(mask1)
        count = 0
        for i in range(1, num_class + 1):
            mask1_i = (mask1 == i).float()
            mask2_i = (mask2 == i).float()
            if torch.any(mask1_i) and torch.any(mask2_i):
                warp_mask1_i = SpatialTransformer(mode='bilinear')(mask1_i.float(), dvf)
                seg_loss = seg_loss + DiceLoss()(warp_mask1_i, mask2_i)
                count = count + 1
        seg_loss = seg_loss / count if count != 0 else seg_loss
    smooth_loss = GradientLoss2D()(dvf)
    total_loss = similarity_loss + args.seg_loss_weight * seg_loss + args.smooth_loss_weight * smooth_loss
    loss_dict = {"total_loss": total_loss, "similarity_loss": similarity_loss, "seg_loss": seg_loss,
                 "smooth_loss": smooth_loss}
    return loss_dict


def create_argparser():
    defaults = dict(
        view_1_dim=2,
        view_2_dim=4,
        data_dir=r"",
        batch_size=16,
        log_interval=10,
        save_interval=50,
        resume_checkpoint="",
        logdir="",
        gpu='0',
        max_save_num=5,
        val_interval=5,
        registration_type=1,
        seg_loss_weight=1,
        smooth_loss_weight=1.,
        consistency_loss_weight=0.5,
        resolution="",
        is_sup=True,
        num_classes=11,
        use_view_1=True
    )
    defaults.update(voxelmorph_defaults_var1())
    parser = argparse.ArgumentParser()
    add_dict_to_argparser(parser, defaults)
    return parser


if __name__ == "__main__":
    main()
