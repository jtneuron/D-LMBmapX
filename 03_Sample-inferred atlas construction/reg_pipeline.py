import os
from argparse import ArgumentParser

import ants
import numpy as np
import SimpleITK as sitk
from tqdm import tqdm
import json
import toml

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def cal_dice(x, y):
    return 2. * (x * y).sum() / (x.sum() + y.sum())

def sitk_read(path):
    return sitk.GetArrayFromImage(sitk.ReadImage(path))

def sitk_write(arr, path):
    sitk.WriteImage(sitk.GetImageFromArray(arr), path)

def reg_pipeline(fixed, moving, type_of_transform, *to_transforms, 
                 **registration_kwargs):
    """
    fixed: 固定图像
    moving: 浮动图像
    type_of_transform: 变换类型
    label: label
    to_transforms: 需要施加形变的图像
    """
    mytx = ants.registration(fixed=fixed,
                             moving=moving,
                             type_of_transform=type_of_transform,
                             **registration_kwargs)  

    registered_img = mytx["warpedmovout"]

    result = [ants.apply_transforms(fixed=to_transform,
                                    moving=to_transform,
                                    transformlist=mytx["fwdtransforms"],
                                    interpolator="nearestNeighbor") 
              for to_transform in to_transforms]
    
    for i in mytx["fwdtransforms"]:
        if os.path.exists(i):
            os.remove(i)
    for i in mytx["invtransforms"]:
        if os.path.exists(i):
            os.remove(i)

    return registered_img, result


def save_as_xlsx_multiple_rows(res, file_path, name, metrics, number=0):
    """
    将形如:
        res = {
            'region1': [val1, val2, val3, ...],
            'region2': [val1, val2, val3, ...]
            ...
        }
    的字典写成 XLSX：
    
    行 1：各 region 分段合并单元格，显示 region 名称
    行 2：各 region 的列，分别显示传入的指标名称
    行 3：与行 2 对应的数值
    支持动态指标数量。
    
    参数:
        res: 字典，包含 region 和对应的指标值
        file_path: 保存的文件路径
        name: 当前数据的名称
        metrics: 指标名称列表，例如 ["Dice", "Affine intensity", "SyN intensity"]
        number: 起始行偏移量
    """
    if not os.path.exists(file_path) or number == 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # 依次写各 region，每个 region 占 len(metrics) 列
        current_index = 0
        ws.cell(row=3, column=1, value=name)
        for region, values in res.items():
            start_col = 2 + current_index * len(metrics)
            end_col = start_col + len(metrics) - 1
            
            # (行 1) 合并单元格，并写 region 名称
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )
            ws.cell(row=1, column=start_col, value=region)
            
            # (行 2) 写指标名称
            for i, metric in enumerate(metrics):
                ws.cell(row=2, column=start_col + i, value=metric)
            
            # (行 3) 写具体数值
            for i, value in enumerate(values):
                ws.cell(row=3, column=start_col + i, value=value)
            
            current_index += 1
    
    else:
        wb = load_workbook(file_path)
        ws = wb.active
        
        st = number + 3
        
        current_index = 0

        ws.cell(row=st, column=1, value=name)
        for region, values in res.items():
            start_col = 2 + current_index * len(metrics)
            
            # (行 3) 写具体数值
            for i, value in enumerate(values):
                ws.cell(row=st, column=start_col + i, value=value)
            
            current_index += 1
        
    # 最后保存 XLSX 到指定目录
    wb.save(file_path)



def main():
    parser = ArgumentParser()
    parser.add_argument('--data_dir',
                        help='dataset config')
    parser.add_argument('--output_dir',
                        help='config path to images')
    parser.add_argument('--mode',
                    default='SyN',
                    choices=["Affine", "SyN"],
                    help="mode of registration")
    args = parser.parse_args()


    # -----读取config-------
    if args.data_dir.endswith('toml'):
        with open(args.data_dir, 'r') as f:
            dataset_config = toml.load(f)
    elif args.data_dir.endswith('json'):
        with open(args.data_dir, 'r') as f:
            dataset_config = json.load(f)
    else:
        raise Exception("dataset config file format error")
        
    moving_paths = dataset_config["moving"]
    registration_type = dataset_config["registration_type"]
    if registration_type == 1:
        fixed_paths = [dataset_config["atlas"]] * len(moving_paths)
    elif registration_type == 4:
        fixed_paths = dataset_config["fixed"]
    else:
        raise ValueError("registration_type err")
    to_transform_suffix = dataset_config.get("to_transform", None)
    save_name = dataset_config.get("save_name", "moving")
    
    for j, moving_path in enumerate(tqdm(moving_paths)):
        moving_name = os.path.basename(moving_path).split(".")[0]
        moving = ants.from_numpy(sitk_read(moving_path))
        moving_label = ants.from_numpy(sitk_read(os.path.join(moving_path.replace(".nii.gz", "_label.nii.gz"))))

        fixed_name = os.path.basename(fixed_paths[j]).split(".")[0]
        fixed = ants.from_numpy(sitk_read(fixed_paths[j]))
        fixed_label = ants.from_numpy(sitk_read(fixed_paths[j].replace(".nii.gz", "_label.nii.gz")))

        if save_name == "fixed":
            name = fixed_name
        else:
            name = moving_name

        if to_transform_suffix is not None:
            to_transform = [ants.from_numpy(sitk_read(os.path.join(moving_path.replace(".nii.gz", f"_{x}.nii.gz"))).astype(np.uint16))
                        for x in to_transform_suffix]
        else:
            to_transform_suffix = []
            to_transform = []

        warped_moving, (warped_label, *warped_transform) = reg_pipeline(fixed, moving, args.mode, moving_label, *to_transform)
        warped_moving = warped_moving.numpy()
        warped_label = warped_label.numpy()
        warped_transform = [x.numpy() for x in warped_transform]

        save_dir = os.path.join(args.output_dir, name)
        os.makedirs(save_dir, exist_ok=True)
        sitk_write(warped_moving.astype(np.uint8), os.path.join(save_dir, f"{name}.nii.gz"))
        sitk_write(warped_label.astype(np.uint8), os.path.join(save_dir, f"{name}_label.nii.gz"))
        for i, suffix in enumerate(to_transform_suffix):
            sitk_write(warped_transform[i], os.path.join(save_dir, f"{name}_{suffix}.nii.gz"))
            
        mx = fixed_label.max().astype(np.int32)
        res = {str(i): [0] for i in range(1, mx + 1)}
        for i in range(1, mx + 1):
            res[str(i)][0] = cal_dice(fixed_label == i, warped_label == i)
        save_as_xlsx_multiple_rows(res, os.path.join(args.output_dir, "res.xlsx"), name, ["Dice"], j)



if __name__ == "__main__":
    main()