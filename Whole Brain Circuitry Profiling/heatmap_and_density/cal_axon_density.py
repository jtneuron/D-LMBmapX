import os
import pandas as pd
import collections
import SimpleITK as sitk
import numpy as np
import json
import copy

from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Border, Side, Font


# a = cal_distribution(sitk.GetArrayFromImage(sitk.ReadImage("/media/user/phz/run_elastix/output/results_12/Slc6a2_P4_20/after_elastix/100040935_139/anno.png")))
# count1 = 0
# count2 = 0
# for k, v in a.items():
#     count1 += v
#     if k not in list(df['id_ytw']):
#         count2 += v
# print(count1, count2)


def col_num_to_letter(col_num):
    letter = ""
    while col_num > 0:
        col_num -= 1
        letter = chr(col_num % 26 + ord('A')) + letter
        col_num //= 26
    return letter


def cal_distribution(img):
    '''
    返回除了背景为0的各个像素值占的面积
    '''
    res = collections.Counter(img.flatten())
    del res[0]
    return res


def write_id_ytw_recur(cfg, id_dict):
    cfg['id_ytw'] = id_dict[cfg['id']]['id_ytw']
    cfg['total_voxels'] = 0
    cfg['seg_voxels'] = 0
    for x in cfg['children']:
        write_id_ytw_recur(x, id_dict)


def update_total_voxels_recur(cfg, distribution):
    if 'total_voxels' not in cfg:
        cfg['total_voxels'] = 0

    if len(cfg['children']):
        for child in cfg['children']:
            cfg['total_voxels'] += update_total_voxels_recur(child, distribution)
    if distribution[cfg['id_ytw']]:
        cfg['total_voxels'] += distribution[cfg['id_ytw']]

    return cfg['total_voxels']


def update_seg_voxels_recur(cfg, distribution):
    if 'seg_voxels' not in cfg:
        cfg['seg_voxels'] = 0

    if len(cfg['children']):
        for child in cfg['children']:
            cfg['seg_voxels'] += update_seg_voxels_recur(child, distribution)

    if distribution[cfg['id_ytw']]:
        cfg['seg_voxels'] += distribution[cfg['id_ytw']]

    return cfg['seg_voxels']


def analyse_statistics(cfg, res=None):
    if res is None:
        keys = ['Brain regions', 'Acronym', 'Number of seg voxels', 'Number of brain region voxels', 'Density']
        deep_copied_dict = {key: [] for key in keys}
        res = {key: copy.deepcopy(deep_copied_dict) for key in range(12)}

    level = cfg['st_level']
    res[level]['Brain regions'].append(cfg['name'])
    res[level]['Acronym'].append(cfg['acronym'])
    res[level]['Number of seg voxels'].append(cfg['seg_voxels'])
    res[level]['Number of brain region voxels'].append(cfg['total_voxels'])
    if cfg['total_voxels'] > 0:
        res[level]['Density'].append(cfg['seg_voxels'] / cfg['total_voxels'])
    else:
        res[level]['Density'].append(0)

    if len(cfg['children']):
        for child in cfg['children']:
            analyse_statistics(child, res)
    return res


def write_to_excel(cfg, file_path):
    '''
    cfg = {
        'level0': {
            'region name': ['a', 'b', 'c', ...],
            'field intensity': [1, 2, 3, ...],
            ...
        }
        ...
    }
    
    file_path 保存位置
    '''
    res = None
    for index, data in cfg.items():
        df = pd.DataFrame(data)
        df.index = range(1, len(df) + 1)

        if res is not None:
            df.insert(0, '', df.index)
            res = pd.concat([res, df], axis=1)

        else:
            df.insert(0, '', df.index)
            res = df

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        res.to_excel(writer, index=False, startrow=1, startcol=0, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']

        # 去除标题行的边框
        for cell in worksheet[2]:
            cell.border = Border(left=Side(border_style=None),
                                 right=Side(border_style=None),
                                 top=Side(border_style=None),
                                 bottom=Side(border_style=None))

        content_font = Font(name='Arial', size=11)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = content_font
                cell.alignment = Alignment(horizontal='left')

        # 合并同一个level的单元格
        for i in range(12):
            a = 6 * i + 2
            b = a + 4
            a = col_num_to_letter(a)
            b = col_num_to_letter(b)
            alignment = Alignment(horizontal='center', vertical='center')
            worksheet.merge_cells(f'{a}1:{b}1')
            worksheet[f'{a}1'] = f'Level-{i}'
            worksheet[f'{a}1'].alignment = alignment

        # 设置列宽
        for col_idx, col in enumerate(worksheet.columns, start=1):
            if col_idx % 6 == 1:
                continue
            for cell in col:
                if isinstance(cell, MergedCell):
                    continue
                column_letter = cell.column_letter
                worksheet.column_dimensions[column_letter].width = 17
                break


def get_density_values(cfg_path, anno_path, seg_path):
    with open(cfg_path, 'r') as f:
        cfg = json.load(f)

    anno = sitk.GetArrayFromImage(sitk.ReadImage(anno_path))
    anno_distribution = cal_distribution(anno)
    update_total_voxels_recur(cfg, anno_distribution)

    seg = sitk.GetArrayFromImage(sitk.ReadImage(seg_path))
    # seg[seg < 5] = 0
    seg[seg != 0] = 1
    res = anno * (seg > 0)
    seg_distribution = cal_distribution(res)

    # res = seg / 255
    # # 初始化一个字典，用于存储每个标签的分割贡献
    # seg_distribution = collections.Counter()
    # # 获取所有唯一标签值（忽略背景，假设背景为0）
    # unique_labels = np.unique(anno)
    # unique_labels = unique_labels[unique_labels != 0]  # 去除背景标签值0
    # # 遍历所有标签值
    # for label in unique_labels:
    #     # 创建一个布尔掩码，提取当前标签区域
    #     label_mask = (anno == label)
    #     # 计算该标签区域内的分割值 B，并累加
    #     label_contribution = np.sum(label_mask * res)  # 计算该区域的分割贡献
    #     # 将该标签的分割像素数量累加到分割分布中
    #     seg_distribution[label] += label_contribution

    update_seg_voxels_recur(cfg, seg_distribution)

    res = analyse_statistics(cfg)
    return res


def get_density_by_acronym(res, brain_region_keys, file_path):
    analysis_dict = {'Acronym': [], 'Number of seg voxels': [], 'Number of brain region voxels': [], 'Density': []}
    # 遍历每个level
    for level in res.keys():
        for brain_region in res[level]['Acronym']:
            if brain_region in brain_region_keys:
                idx = res[level]['Acronym'].index(brain_region)
                analysis_dict['Acronym'].append(brain_region)
                analysis_dict['Number of seg voxels'].append(res[level]['Number of seg voxels'][idx])
                analysis_dict['Number of brain region voxels'].append(res[level]['Number of brain region voxels'][idx])
                analysis_dict['Density'].append(res[level]['Density'][idx])

    # diff = [item for item in brain_region_keys if item not in temp]
    PRNc_index = analysis_dict['Acronym'].index('PRNc')
    PRNr_index = analysis_dict['Acronym'].index('PRNr')
    # print(PRNc_index, PRNr_index)
    analysis_dict['Acronym'].append("PRN")
    PRN_seg_voxels = analysis_dict['Number of seg voxels'][PRNc_index] + analysis_dict['Number of seg voxels'][
        PRNr_index]
    analysis_dict['Number of seg voxels'].append(PRN_seg_voxels)
    PRN_brain_region_voxels = analysis_dict['Number of brain region voxels'][PRNc_index] + \
                              analysis_dict['Number of brain region voxels'][PRNr_index]
    analysis_dict['Number of brain region voxels'].append(PRN_brain_region_voxels)
    analysis_dict['Density'].append(PRN_seg_voxels / PRN_brain_region_voxels)

    for key in analysis_dict:
        analysis_dict[key].pop(PRNc_index)
        analysis_dict[key].pop(PRNr_index - 1)

    # write to excel
    df = pd.DataFrame(analysis_dict)
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Analysis')
        worksheet = writer.sheets['Analysis']

        # 设置字体和对齐
        content_font = Font(name='Arial', size=11)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = content_font
                cell.alignment = Alignment(horizontal='left')

        # 去除标题行的边框
        for cell in worksheet[1]:
            cell.border = Border(left=Side(border_style=None),
                                 right=Side(border_style=None),
                                 top=Side(border_style=None),
                                 bottom=Side(border_style=None))


if __name__ == "__main__":

    cfg_path = "add_id_ytw.json"
    
    anno_path = r"P28_anno.tiff"
    excel_output_path = r"axon_density.xlsx"
    analysis_excel = r"axon_density_special_brain_region_analysis.xlsx"
    seg_path1 = r"P28_axon_final.tiff"
    
    brain_region_keys = ['ADP', 'ACB', 'ARH', 'ASO', 'BMA', 'BST', 'CEA', 'CP', 'DMH', 'FS', 'GPi', 'IA', 'IF', 'LC',
                         'LGv', 'LPO', 'LRN', 'LS', 'MDRNd', 'ME', 'MEA', 'MOB', 'NLOT', 'NTS', 'OT', 'PAG', 'PD',
                         'PGRNl', 'PH', 'PRNr', 'PRNc', 'PS', 'PVH', 'RR', 'SCH', 'SNc', 'SNr', 'SOC', 'STN', 'VTA', 'ZI', 'SUM',
                         'RE', 'Isocortex', 'ACA', 'ORB', 'LH']
    
    res = get_density_values(cfg_path, anno_path, seg_path1)
    write_to_excel(res, excel_output_path)

    Fetch brain regions for analysis
    get_density_by_acronym(res, brain_region_keys, analysis_excel)

